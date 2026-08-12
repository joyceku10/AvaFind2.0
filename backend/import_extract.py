"""Import a demand extract (.xlsx) into the AvaFind database.

Usage:
    python import_extract.py <path-to-extract.xlsx> [--as-of YYYY-MM-DD] [--sheet NAME]

The target database comes from the same env vars the API uses (see app/db.py):
local SQLite by default, Azure SQL when SQL_SERVER/... are set.

The import is a full replace: existing rows are deleted and the file's rows
inserted in one transaction. The extract date is taken from --as-of, or from a
leading YYYYMMDD in the filename (e.g. 20260211-A&I_US_Demand.xlsx).
"""

import argparse
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl

from app.db import SessionLocal, engine
from app.models import Base, ImportMeta, Role, RoleSkill

# Excel header -> Role attribute, 1:1 with the extract. If a future extract
# renames or drops a column, the import fails loudly instead of guessing.
COLUMNS = {
    "RFE 8": "rfe8",
    "RFE 9": "rfe9",
    "Client": "client",
    "Project ID": "project_id",
    "Project": "project_name",
    "Role Primary Contact": "primary_contact",
    "Role ID": "role_id",
    "Role Title": "role_title",
    "Assigned Role": "assigned_role",
    "Role Fulfillment Contact": "fulfillment_contact",
    "Role Job Family Group": "job_family_group",
    "Role Status": "role_status",
    "Sold Role": "sold_role",
    "Charg Role": "charg_role",
    "Channel": "channel",
    "Min Role Level": "min_level",
    "Max Role Level": "max_level",
    "Role Start Date": "start_date",
    "Role End Date": "end_date",
    "Primary Skill": None,  # parsed into primary_skill_name/_proficiency
    "Role Skills": "skills_raw",
    "Project Geo": "project_geo",
    "Role Work Location": "work_location",
    "Role Priority": "priority",
    "Role Created Date": "created_date",
    "Role Description": "description",
}

BOOL_FIELDS = {"sold_role", "charg_role"}
INT_FIELDS = {"min_level", "max_level"}
DATE_FIELDS = {"start_date", "end_date", "created_date"}

# One packed skill entry: "3 - Skill Name   (P2 - Intermediate)". The
# proficiency suffix is optional and matched strictly as "(P<n> - Word)" so
# parentheses inside skill names (e.g. "... Business Intelligence (BI)") are
# kept as part of the name.
SKILL_PROFICIENCY_RE = re.compile(r"\(\s*(P\d\s*-\s*[A-Za-z]+)\s*\)\s*$")
SKILL_PREFIX_RE = re.compile(r"^\s*\d+\s*-\s*")


def parse_skill_entry(entry: str):
    entry = entry.strip()
    if not entry:
        return None
    proficiency = None
    m = SKILL_PROFICIENCY_RE.search(entry)
    if m:
        proficiency = re.sub(r"\s+", " ", m.group(1))
        entry = entry[: m.start()]
    name = SKILL_PREFIX_RE.sub("", entry)
    name = re.sub(r"\s+", " ", name).strip()
    if not name:
        return None
    return name, proficiency


def parse_skills(raw: str):
    return [s for part in raw.split("|") if (s := parse_skill_entry(part))]


def clean(value):
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def to_date(value, field, role_id, warnings):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    warnings.append(f"role {role_id}: unparseable {field}: {value!r}")
    return None


def extract_date_from_filename(path: Path):
    m = re.match(r"^(\d{8})", path.name)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(1), "%Y%m%d").date()
    except ValueError:
        return None


def read_rows(path: Path, sheet_name: str | None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [clean(h) for h in next(rows)]
    missing = [h for h in COLUMNS if h not in headers]
    if missing:
        sys.exit(f"ERROR: expected columns missing from {ws.title!r}: {missing}\n"
                 f"Found: {headers}")
    extra = [h for h in headers if h and h not in COLUMNS]
    if extra:
        print(f"NOTE: ignoring unrecognized columns: {extra}")
    index = {h: headers.index(h) for h in COLUMNS}
    return ws.title, index, [r for r in rows if any(v is not None and str(v).strip() for v in r)]


def build_role(raw_row, index, warnings):
    get = lambda header: clean(raw_row[index[header]])
    role_id = get("Role ID")
    if not role_id:
        warnings.append(f"row skipped: no Role ID ({raw_row[:5]!r}...)")
        return None, []

    values = {}
    for header, attr in COLUMNS.items():
        if attr is None or attr == "role_id":
            continue
        v = get(header)
        if attr in BOOL_FIELDS and v is not None:
            v = str(v).strip().lower() == "yes"
        elif attr in INT_FIELDS and v is not None:
            try:
                v = int(v)
            except ValueError:
                warnings.append(f"role {role_id}: non-integer {header}: {v!r}")
                v = None
        elif attr in DATE_FIELDS:
            v = to_date(v, header, role_id, warnings)
        values[attr] = v

    role = Role(role_id=str(role_id), **values)

    primary = get("Primary Skill")
    if primary:
        parsed = parse_skill_entry(primary)
        if parsed:
            role.primary_skill_name, role.primary_skill_proficiency = parsed
        else:
            warnings.append(f"role {role_id}: unparseable Primary Skill: {primary!r}")

    skills = []
    if role.skills_raw:
        entries = parse_skills(role.skills_raw)
        if not entries:
            warnings.append(f"role {role_id}: no skills parsed from: {role.skills_raw!r}")
        skills = [RoleSkill(position=i + 1, name=n, proficiency=p)
                  for i, (n, p) in enumerate(entries)]
    return role, skills


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("file", type=Path)
    ap.add_argument("--as-of", type=lambda s: datetime.strptime(s, "%Y-%m-%d").date(),
                    help="extract date; defaults to leading YYYYMMDD in the filename")
    ap.add_argument("--sheet", help="sheet name; defaults to the first sheet")
    args = ap.parse_args()

    if not args.file.exists():
        sys.exit(f"ERROR: file not found: {args.file}")

    extract_date = args.as_of or extract_date_from_filename(args.file)
    if extract_date is None:
        sys.exit("ERROR: could not derive extract date from filename; pass --as-of YYYY-MM-DD")

    sheet, index, raw_rows = read_rows(args.file, args.sheet)
    print(f"Read {len(raw_rows)} data rows from sheet {sheet!r}")

    warnings: list[str] = []
    roles, all_skills, seen = [], [], set()
    for raw in raw_rows:
        role, skills = build_role(raw, index, warnings)
        if role is None:
            continue
        if role.role_id in seen:
            warnings.append(f"duplicate Role ID {role.role_id}: keeping first, skipping later row")
            continue
        seen.add(role.role_id)
        role.skills = skills
        roles.append(role)
        all_skills.extend(skills)

    Base.metadata.create_all(engine)
    with SessionLocal() as session:
        session.query(RoleSkill).delete()
        session.query(Role).delete()
        session.query(ImportMeta).delete()
        session.add_all(roles)
        session.add(ImportMeta(
            id=1,
            source_file=args.file.name,
            extract_date=extract_date,
            imported_at=datetime.now(timezone.utc),
            row_count=len(roles),
        ))
        session.commit()

    print(f"Imported {len(roles)} roles, {len(all_skills)} skill rows "
          f"(extract date {extract_date.isoformat()})")
    if warnings:
        print(f"\n{len(warnings)} warning(s):")
        for w in warnings:
            print(f"  - {w}")
    else:
        print("No warnings.")


if __name__ == "__main__":
    main()
